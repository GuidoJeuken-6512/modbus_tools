import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pymodbus.client import ModbusTcpClient
import csv
import threading
import time
import logging
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# Scanner configuration constants
DEFAULT_START_REGISTER = 0
DEFAULT_STOP_REGISTER = 10000
LARGE_RANGE_WARNING_THRESHOLD = 10000

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S'
)
logger = logging.getLogger(__name__)


class ModbusScanner:
    def __init__(self, root):
        self.root = root
        self.root.title("Modbus Register Scanner")
        self.root.geometry("800x600")

        # Scanner state
        self.scanning = False
        self.scan_thread = None
        self.found_registers = []

        # Create GUI
        self.create_widgets()

    def create_widgets(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Configuration frame
        config_frame = ttk.LabelFrame(
            main_frame, text="Scanner Configuration", padding="10"
        )
        config_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E),
                          pady=(0, 10))

        # Server IP
        ttk.Label(config_frame, text="Server IP:").grid(
            row=0, column=0, padx=5, pady=5, sticky="e"
        )
        self.ip_entry = ttk.Entry(config_frame, width=20)
        self.ip_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        self.ip_entry.insert(0, "192.168.178.125")

        # Port
        ttk.Label(config_frame, text="Port:").grid(
            row=0, column=2, padx=5, pady=5, sticky="e"
        )
        self.port_entry = ttk.Entry(config_frame, width=10)
        self.port_entry.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.port_entry.insert(0, "502")

        # Register range
        ttk.Label(config_frame, text="Start Register:").grid(
            row=1, column=0, padx=5, pady=5, sticky="e"
        )
        self.start_entry = ttk.Entry(config_frame, width=10)
        self.start_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
        self.start_entry.insert(0, str(DEFAULT_START_REGISTER))

        ttk.Label(config_frame, text="Stop Register:").grid(
            row=1, column=2, padx=5, pady=5, sticky="e"
        )
        self.stop_entry = ttk.Entry(config_frame, width=10)
        self.stop_entry.grid(row=1, column=3, padx=5, pady=5, sticky="w")
        self.stop_entry.insert(0, str(DEFAULT_STOP_REGISTER))

        # Register type
        ttk.Label(config_frame, text="Register Type:").grid(
            row=2, column=0, padx=5, pady=5, sticky="e"
        )
        self.register_type = ttk.Combobox(
            config_frame, values=["Holding Registers", "Input Registers"],
            width=15
        )
        self.register_type.grid(row=2, column=1, padx=5, pady=5, sticky="w")
        self.register_type.set("Holding Registers")

        # Scan delay
        ttk.Label(config_frame, text="Delay (ms):").grid(
            row=2, column=2, padx=5, pady=5, sticky="e"
        )
        self.delay_entry = ttk.Entry(config_frame, width=10)
        self.delay_entry.grid(row=2, column=3, padx=5, pady=5, sticky="w")
        self.delay_entry.insert(0, "50")

        # Retry settings
        ttk.Label(config_frame, text="Max Retries:").grid(
            row=3, column=0, padx=5, pady=5, sticky="e"
        )
        self.retry_entry = ttk.Entry(config_frame, width=10)
        self.retry_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")
        self.retry_entry.insert(0, "3")

        ttk.Label(config_frame, text="Timeout (s):").grid(
            row=3, column=2, padx=5, pady=5, sticky="e"
        )
        self.timeout_entry = ttk.Entry(config_frame, width=10)
        self.timeout_entry.grid(row=3, column=3, padx=5, pady=5, sticky="w")
        self.timeout_entry.insert(0, "5")

        # Control buttons
        button_frame = ttk.Frame(config_frame)
        button_frame.grid(row=4, column=0, columnspan=4, pady=10)

        self.scan_button = ttk.Button(
            button_frame, text="Start Scan", command=self.start_scan
        )
        self.scan_button.pack(side=tk.LEFT, padx=5)

        self.stop_button = ttk.Button(
            button_frame, text="Stop Scan", command=self.stop_scan,
            state=tk.DISABLED
        )
        self.stop_button.pack(side=tk.LEFT, padx=5)

        self.export_csv_button = ttk.Button(
            button_frame, text="Export CSV", command=self.export_csv,
            state=tk.DISABLED
        )
        self.export_csv_button.pack(side=tk.LEFT, padx=5)

        self.export_excel_button = ttk.Button(
            button_frame, text="Export Excel", command=self.export_excel,
            state=tk.DISABLED
        )
        self.export_excel_button.pack(side=tk.LEFT, padx=5)

        self.clear_button = ttk.Button(
            button_frame, text="Clear Results", command=self.clear_results
        )
        self.clear_button.pack(side=tk.LEFT, padx=5)

        # Progress bar
        self.progress = ttk.Progressbar(main_frame, mode='determinate')
        self.progress.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E),
                           pady=5)

        # Status label
        self.status_label = ttk.Label(main_frame, text="Ready to scan")
        self.status_label.grid(row=2, column=0, columnspan=1, pady=5,
                               sticky="w")

        # Connection status label
        self.connection_status = ttk.Label(main_frame, text="",
                                           foreground="green")
        self.connection_status.grid(row=2, column=1, columnspan=1, pady=5,
                                    sticky="e")

        # Results frame
        results_frame = ttk.LabelFrame(main_frame, text="Scan Results",
                                       padding="10")
        results_frame.grid(row=3, column=0, columnspan=2,
                           sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))

        # Treeview for results
        columns = ("Register", "Value (Hex)", "Value (Dec)", "Type",
                   "Raw Data")
        self.tree = ttk.Treeview(results_frame, columns=columns,
                                 show="headings", height=15)

        # Configure columns
        self.tree.heading("Register", text="Register")
        self.tree.heading("Value (Hex)", text="Value (Hex)")
        self.tree.heading("Value (Dec)", text="Value (Dec)")
        self.tree.heading("Type", text="Type")
        self.tree.heading("Raw Data", text="Raw Data")

        self.tree.column("Register", width=80)
        self.tree.column("Value (Hex)", width=140)
        self.tree.column("Value (Dec)", width=120)
        self.tree.column("Type", width=180)
        self.tree.column("Raw Data", width=150)

        # Scrollbar for treeview
        scrollbar = ttk.Scrollbar(results_frame, orient=tk.VERTICAL,
                                  command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))

        # Configure grid weights
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(3, weight=1)
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

    def start_scan(self):
        """Start the scanning process"""
        try:
            start_reg = int(self.start_entry.get())
            stop_reg = int(self.stop_entry.get())
            delay = int(self.delay_entry.get())
        except ValueError:
            messagebox.showerror("Input Error", "Please enter valid numeric "
                                                "values.")
            return

        if start_reg >= stop_reg:
            messagebox.showerror("Input Error", "Start register must be less "
                                                "than stop register.")
            return

        if stop_reg - start_reg > LARGE_RANGE_WARNING_THRESHOLD:
            messagebox.showwarning("Warning", "Large range selected. This may "
                                              "take a long time.")

        self.scanning = True
        self.scan_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        self.export_csv_button.config(state=tk.DISABLED)
        self.export_excel_button.config(state=tk.DISABLED)

        # Clear previous results
        self.clear_results()

        # Start scan in separate thread
        self.scan_thread = threading.Thread(
            target=self.scan_registers, args=(start_reg, stop_reg, delay)
        )
        self.scan_thread.daemon = True
        self.scan_thread.start()

    def stop_scan(self):
        """Stop the scanning process"""
        self.scanning = False
        self.scan_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.status_label.config(text="Scan stopped by user")

    def create_client(self, ip, port):
        """Create and connect Modbus client with retry logic"""
        max_retries = 3
        retry_delay = 1  # seconds
        timeout = 5
        
        try:
            timeout = int(self.timeout_entry.get())
        except ValueError:
            timeout = 5
        
        for attempt in range(max_retries):
            try:
                client = ModbusTcpClient(ip, port=port, timeout=timeout)
                if client.connect():
                    logger.info(f"Connected to {ip}:{port} (attempt {attempt + 1})")
                    return client
                else:
                    logger.warning(f"Connection attempt {attempt + 1} failed")
                    client.close()
            except Exception as e:
                logger.warning(f"Connection attempt {attempt + 1} error: {e}")
            
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
                retry_delay *= 2  # Exponential backoff
        
        return None

    def read_register_with_retry(self, client, register, reg_type,
                                 max_retries=None):
        """Read a single register with retry logic and auto-reconnect
        Returns:
        - dict with register data if successful
        - None if register not found (normal for scanning)
        - 'CONNECTION_ERROR' if connection failed
        """
        if max_retries is None:
            try:
                max_retries = int(self.retry_entry.get())
            except ValueError:
                max_retries = 3
                
        for attempt in range(max_retries):
            try:
                if reg_type == "Holding Registers":
                    response = client.read_holding_registers(
                        address=register, count=1)
                else:
                    response = client.read_input_registers(
                        address=register, count=1)

                if hasattr(response, 'registers') and response.registers:
                    # Try to read 2 registers to check if this might be a 32-bit register
                    # Only do this if the single register read was successful
                    try:
                        if reg_type == "Holding Registers":
                            response_32 = client.read_holding_registers(
                                address=register, count=2)
                        else:
                            response_32 = client.read_input_registers(
                                address=register, count=2)
                            
                        if (hasattr(response_32, 'registers') and
                            len(response_32.registers) == 2):
                            # Check if the second register also has a meaningful value
                            # or if the first register has a very large value (indicating 32-bit)
                            first_reg = response_32.registers[0]
                            second_reg = response_32.registers[1]
                            
                            # Heuristic: if first register > 32767 or second register > 0,
                            # it's likely a 32-bit register
                            if (first_reg > 32767 or second_reg > 0 or
                                first_reg == 0 and second_reg > 0):
                                # Combine two 16-bit values to 32-bit (big-endian)
                                value_32 = ((first_reg << 16) | second_reg)
                                return {
                                    'value': value_32,
                                    'is_32bit': True,
                                    'raw_registers': response_32.registers
                                }
                    except Exception:
                        # If 2-register read fails, fall back to single register
                        pass
                    
                    # Return as 16-bit register (either unknown or fallback)
                    return {
                        'value': response.registers[0],
                        'is_32bit': False,
                        'raw_registers': response.registers
                    }
                else:
                    return None

            except Exception as exc:
                error_str = str(exc).lower()
                
                # Check if it's a Modbus protocol error (not a connection error)
                if any(keyword in error_str for keyword in
                       ["illegal data address", "illegal function",
                        "illegal data value", "slave device failure",
                        "acknowledge", "slave device busy",
                        "modbus exception", "exception code"]):
                    # This is a Modbus protocol error, not a connection error
                    # Don't retry or reconnect for these errors - they're expected
                    logger.debug(f"Modbus protocol error for register "
                                 f"{register}: {exc}")
                    return None

                # Check if it's a connection error
                elif any(keyword in error_str for keyword in
                         ["10054", "connection", "reset", "closed", "timeout",
                          "network", "connection refused", "connection aborted",
                          "broken pipe"]):
                    logger.warning(f"Connection error reading register "
                                   f"{register} (attempt {attempt + 1}): {exc}")
                    logger.info("Connection lost, attempting to reconnect...")
                    client.close()

                    # Try to reconnect
                    ip = self.ip_entry.get()
                    port = int(self.port_entry.get())
                    new_client = self.create_client(ip, port)
                    if new_client:
                        # Update the client reference
                        client = new_client
                        logger.info("Reconnected successfully")
                    else:
                        logger.error("Failed to reconnect")
                        return 'CONNECTION_ERROR'
                else:
                    # For other errors, just retry with exponential backoff
                    logger.warning(f"Error reading register {register} "
                                   f"(attempt {attempt + 1}): {exc}")
                    if attempt < max_retries - 1:
                        sleep_time = 0.5 * (2 ** attempt)  # Exponential backoff
                        time.sleep(sleep_time)
        
        return None

    def scan_registers(self, start_reg, stop_reg, delay):
        """Scan registers in a separate thread with robust connection handling"""
        ip = self.ip_entry.get()
        port = int(self.port_entry.get())
        reg_type = self.register_type.get()

        # Create initial connection
        self.root.after(0, lambda: self.connection_status.config(
            text="Connecting...", foreground="orange"))
        client = self.create_client(ip, port)
        if not client:
            self.root.after(0, lambda: messagebox.showerror(
                "Connection Error", f"Unable to connect to {ip}:{port}"))
            self.root.after(0, lambda: self.connection_status.config(
                text="Connection Failed", foreground="red"))
            self.root.after(0, self.stop_scan)
            return
        
        self.root.after(0, lambda: self.connection_status.config(
            text="Connected", foreground="green"))

        try:
            total_registers = stop_reg - start_reg + 1
            found_count = 0
            consecutive_errors = 0
            max_consecutive_errors = 10

            for i, register in enumerate(range(start_reg, stop_reg + 1)):
                if not self.scanning:
                    break

                # Read register with retry logic
                register_data = self.read_register_with_retry(client, register, reg_type)
                
                if register_data == 'CONNECTION_ERROR':
                    # Connection error - count as consecutive error
                    consecutive_errors += 1
                    
                    # If too many consecutive connection errors, try to reconnect
                    if consecutive_errors >= max_consecutive_errors:
                        logger.warning("Too many consecutive connection errors, "
                                       "attempting to reconnect...")
                        self.root.after(0, lambda: self.connection_status.config(
                            text="Reconnecting...", foreground="orange"))
                        client.close()
                        client = self.create_client(ip, port)
                        if not client:
                            logger.error("Failed to reconnect, stopping scan")
                            self.root.after(0, lambda: self.connection_status.config(
                                text="Connection Lost", foreground="red"))
                            break
                        else:
                            self.root.after(0, lambda: self.connection_status.config(
                                text="Reconnected", foreground="green"))
                        consecutive_errors = 0
                elif register_data is not None:
                    found_count += 1
                    consecutive_errors = 0  # Reset error counter

                    # Extract data from register_data
                    if isinstance(register_data, dict):
                        value = register_data['value']
                        is_32bit = register_data['is_32bit']
                        raw_registers = register_data['raw_registers']
                    else:
                        # Fallback for old format
                        value = register_data
                        is_32bit = False
                        raw_registers = [value]

                    # Format display values
                    if is_32bit:
                        value_hex = f"0x{value:08X}"
                        register_type_display = f"{reg_type} (32-bit)"
                        raw_display = (f"[0x{raw_registers[0]:04X}, "
                                       f"0x{raw_registers[1]:04X}]")
                    else:
                        value_hex = f"0x{value:04X}"
                        register_type_display = f"{reg_type} (16-bit)"
                        raw_display = f"0x{raw_registers[0]:04X}"

                    # Add to results
                    result = {
                        'register': register,
                        'value_hex': value_hex,
                        'value_dec': value,
                        'type': register_type_display,
                        'is_32bit': is_32bit,
                        'raw_registers': raw_registers,
                        'raw_display': raw_display
                    }
                    self.found_registers.append(result)

                    # Update GUI in main thread
                    self.root.after(0, lambda r=result:
                                   self.add_result_to_tree(r))

                    logger.info(f"Found register {register}: {value_hex} "
                                 f"({value}) - {register_type_display} "
                                 f"{raw_display}")
                else:
                    # Register not found - this is normal for scanning
                    # Don't count as error, just continue
                    pass

                # Update progress
                progress = int((i + 1) / total_registers * 100)
                self.root.after(0, lambda p=progress:
                                  self.progress.config(value=p))
                self.root.after(0, lambda i=i, t=total_registers,
                                f=found_count:
                                self.status_label.config(
                                    text=f"Scanning... {i+1}/{t} "
                                         f"registers, {f} found"))

                # Delay between reads
                if delay > 0:
                    time.sleep(delay / 1000.0)

            # Scan completed
            self.root.after(0, lambda: self.status_label.config(
                text=f"Scan completed. Found {found_count} active "
                     f"registers."))
            self.root.after(0, self.stop_scan)
            if found_count > 0:
                self.root.after(0, lambda: self.export_csv_button.config(
                    state=tk.NORMAL))
                self.root.after(0, lambda: self.export_excel_button.config(
                    state=tk.NORMAL))

        except Exception as exc:
            error_msg = f"Scan error: {str(exc)}"
            self.root.after(0, lambda: messagebox.showerror("Error", error_msg))
            self.root.after(0, self.stop_scan)
        finally:
            if client:
                client.close()

    def add_result_to_tree(self, result):
        """Add a result to the treeview"""
        self.tree.insert("", "end", values=(
            result['register'],
            result['value_hex'],
            result['value_dec'],
            result['type'],
            result.get('raw_display', '')
        ))

    def clear_results(self):
        """Clear all results"""
        self.tree.delete(*self.tree.get_children())
        self.found_registers.clear()
        self.progress.config(value=0)
        self.status_label.config(text="Ready to scan")
        self.connection_status.config(text="", foreground="green")

    def export_csv(self):
        """Export results to CSV file"""
        if not self.found_registers:
            messagebox.showwarning("No Data", "No results to export.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Save scan results"
        )

        if filename:
            try:
                with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(['Register', 'Value (Hex)', 'Value (Dec)',
                                     'Type', 'Raw Data', 'Is 32-bit',
                                     'Timestamp'])

                    for result in self.found_registers:
                        writer.writerow([
                            result['register'],
                            result['value_hex'],
                            result['value_dec'],
                            result['type'],
                            result.get('raw_display', ''),
                            'Yes' if result.get('is_32bit', False) else 'No',
                            datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                        ])

                messagebox.showinfo("Export Complete",
                                     f"Results exported to {filename}")
                logger.info(f"Exported {len(self.found_registers)} "
                             f"registers to {filename}")

            except Exception as exc:
                messagebox.showerror("Export Error",
                                     f"Failed to export CSV: {str(exc)}")

    def export_excel(self):
        """Export results to Excel file"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Excel Export Error",
                                 "openpyxl library not installed. "
                                 "Please install it with: pip install openpyxl")
            return

        if not self.found_registers:
            messagebox.showwarning("No Data", "No results to export.")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Save scan results as Excel"
        )

        if filename:
            try:
                wb = Workbook()
                ws = wb.active
                ws.title = "Modbus Scan Results"

                # Define headers
                headers = ['Register', 'Value (Hex)', 'Value (Dec)',
                           'Type', 'Raw Data', 'Is 32-bit', 'Timestamp']
                
                # Style for headers
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="366092",
                                          end_color="366092",
                                          fill_type="solid")
                header_alignment = Alignment(horizontal="center",
                                            vertical="center")

                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # Write data
                for row, result in enumerate(self.found_registers, 2):
                    ws.cell(row=row, column=1, value=result['register'])
                    ws.cell(row=row, column=2, value=result['value_hex'])
                    ws.cell(row=row, column=3, value=result['value_dec'])
                    ws.cell(row=row, column=4, value=result['type'])
                    ws.cell(row=row, column=5, 
                            value=result.get('raw_display', ''))
                    ws.cell(row=row, column=6,
                            value='Yes' if result.get('is_32bit', False) else 'No')
                    ws.cell(row=row, column=7,
                            value=datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except Exception:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # Freeze header row
                ws.freeze_panes = "A2"

                wb.save(filename)
                messagebox.showinfo("Export Complete",
                                     f"Results exported to {filename}")
                logger.info(f"Exported {len(self.found_registers)} "
                             f"registers to {filename}")

            except Exception as exc:
                messagebox.showerror("Export Error",
                                     f"Failed to export Excel: {str(exc)}")


def main():
    root = tk.Tk()
    ModbusScanner(root)
    root.mainloop()


if __name__ == "__main__":
    main()